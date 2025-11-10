# engobras/views.py

from datetime import datetime, timedelta, date
from collections import defaultdict, Counter
import datetime as dt

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db import transaction
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Prefetch, Count
from django.utils.html import escape
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ===== modelos/forms do app ENGOBRAS =====
from .forms import ActivityForm, CollaboratorSettingsForm
from .models import (
    Workday, Activity, ActivitySession,
    Project, Milestone, GeneralActivity, SpecificActivity,
)

# ===== colaborador centralizado & EPE =====
from common.models import Collaborator, Department
from epe.models import PanelSize


# =============================================================================
# Helpers globais
# =============================================================================

AUTO_DEPT_SLUG = "engobras"   # slug do departamento ENGOBRAS (em common.Department)
AUTO_NS = "engobras"          # namespace das URLs do app

def _get_user_collaborator(user):
    """Retorna o Collaborator central do usuário para o departamento ENGOBRAS."""
    if not user or not getattr(user, "is_authenticated", False):
        return None
    return (Collaborator.objects
            .select_related("department", "user")
            .filter(user=user, is_active=True, department__slug=AUTO_DEPT_SLUG)
            .first())

def _require_auto_collaborator(request):
    """Exige que o usuário tenha perfil no depto; senão 403."""
    collab = _get_user_collaborator(request.user)
    if not collab:
        return None, HttpResponseForbidden("Você não possui perfil neste departamento.")
    return collab, None

def get_user_activity_or_404(user, pk):
    """Busca Activity garantindo que pertence ao colaborador do usuário atual (ENGOBRAS)."""
    collab = _get_user_collaborator(user)
    return get_object_or_404(Activity, pk=pk, collaborator=collab)

def _parse_date(s: str):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def _overlap_seconds(start_a, end_a, start_b_date, end_b_date):
    """Segundos de sobreposição entre [start_a, end_a] e o(s) dia(s) [start_b_date..end_b_date]."""
    if end_a is None:
        end_a = timezone.now()
    start_b = timezone.make_aware(datetime.combine(start_b_date, datetime.min.time()))
    end_b   = timezone.make_aware(datetime.combine(end_b_date, datetime.max.time()))
    start = max(start_a, start_b)
    end   = min(end_a, end_b)
    delta = (end - start).total_seconds()
    return int(delta) if delta > 0 else 0

def _format_hms(total_seconds: int) -> str:
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

def _user_is_manager(user) -> bool:
    try:
        c = _get_user_collaborator(user)
        return bool(c and c.is_manager)
    except Exception:
        return False


# =============================================================================
# Dashboard / expediente
# =============================================================================

@login_required
def dashboard(request):
    collab = _get_user_collaborator(request.user)

    # Data selecionada via query (?date=YYYY-MM-DD)
    today = timezone.localdate()
    date_str = request.GET.get("date", "")
    try:
        selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else today
    except ValueError:
        selected_date = today
    is_today = (selected_date == today)

    workday = None
    activities = Activity.objects.none()
    stats = {"total_secs": 0, "active": 0, "paused": 0, "done": 0}

    if collab:
        workday = Workday.objects.filter(collaborator=collab, date=selected_date).first()
        if workday:
            activities = (
                Activity.objects
                .filter(collaborator=collab, workday=workday)
                .select_related("project", "milestone", "general", "specific")
                .prefetch_related("sessions")
                .order_by("-id")
            )
            stats["total_secs"] = sum(a.total_active_seconds() for a in activities)
            stats["active"] = activities.filter(is_active=True, finished_at__isnull=True).count()
            stats["paused"] = activities.filter(is_active=False, finished_at__isnull=True).count()
            stats["done"]   = activities.filter(finished_at__isnull=False).count()

    # prev/next (não permite passar de hoje)
    prev_date = selected_date - timedelta(days=1)
    next_date = selected_date + timedelta(days=1)
    can_next = next_date <= today

    # faixa de dias (pílulas)
    window = []
    span = range(-3, 4)  # 7 dias
    existing_dates = set()
    if collab:
        dates_in_window = [
            selected_date + timedelta(days=i)
            for i in span if (selected_date + timedelta(days=i)) <= today
        ]
        existing_dates = set(
            Workday.objects.filter(
                collaborator=collab, date__in=dates_in_window
            ).values_list("date", flat=True)
        )
    for i in span:
        d = selected_date + timedelta(days=i)
        if d > today:
            continue
        window.append({
            "date": d,
            "iso": d.isoformat(),
            "label": d.strftime("%d/%m"),
            "is_selected": d == selected_date,
            "is_today": d == today,
            "has_workday": d in existing_dates,
        })

    context = {
        "collab": collab,
        "workday": workday,
        "activities": activities,
        "stats": stats,
        "projects": Project.objects.order_by("name"),
        "milestones": Milestone.objects.filter(is_active=True).order_by("order", "name"),
        "selected_date": selected_date,
        "is_today": is_today,
        "prev_date": prev_date,
        "next_date": next_date if can_next else None,
        "nav_days": window,
    }
    return render(request, "engobras/obras/dashboard.html", context)



@login_required
@require_POST
def start_workday(request):
    collab, err = _require_auto_collaborator(request)
    if err:
        return err
    wd, created = Workday.objects.get_or_create(
        collaborator=collab,
        date=timezone.localdate(),
        defaults={"is_open": True},
    )
    if not created and not wd.is_open:
        wd.is_open, wd.ended_at = True, None
        wd.save(update_fields=["is_open", "ended_at"])
    return redirect(f"{AUTO_NS}:dashboard")


# =============================================================================
# Atividades
# =============================================================================

@login_required
@transaction.atomic
def activity_create(request):
    """
    Cria a atividade e já inicia a primeira sessão (cronômetro).
    O ActivityForm deve popular os querysets (Milestone/General/Specific) no __init__.
    """
    collab, err = _require_auto_collaborator(request)
    if err:
        return err

    wd, _ = Workday.objects.get_or_create(
        collaborator=collab,
        date=timezone.localdate(),
        defaults={"is_open": True},
    )

    form = ActivityForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        act: Activity = form.save(commit=False)
        act.collaborator = collab
        act.workday = wd
        act.save()
        # inicia cronômetro
        ActivitySession.objects.create(activity=act)
        return redirect(f"{AUTO_NS}:dashboard")

    return render(request, "engobras/obras/activity_form.html", {"form": form})



@login_required
@require_POST
def activity_start(request, pk):
    act = get_user_activity_or_404(request.user, pk)
    last = act.sessions.order_by("-id").first()
    if not last or last.ended_at:
        ActivitySession.objects.create(activity=act)
    act.is_active = True
    act.save(update_fields=["is_active"])
    return redirect(f"{AUTO_NS}:dashboard")


@login_required
@require_POST
def activity_pause(request, pk):
    act = get_user_activity_or_404(request.user, pk)
    last = act.sessions.order_by("-id").first()
    if last and not last.ended_at:
        last.stop()
    act.is_active = False
    act.save(update_fields=["is_active"])
    return redirect(f"{AUTO_NS}:dashboard")


@login_required
@require_POST
def activity_resume(request, pk):
    return activity_start(request, pk)


@login_required
@require_POST
def activity_finish(request, pk):
    act = get_user_activity_or_404(request.user, pk)
    for s in act.sessions.filter(ended_at__isnull=True):
        s.stop()
    act.is_active = False
    act.finished_at = timezone.now()
    act.save(update_fields=["is_active", "finished_at"])
    return redirect(f"{AUTO_NS}:dashboard")


# =============================================================================
# Páginas simples
# =============================================================================

def profile(request):
    return HttpResponse("Perfil (ENGOBRAS) — placeholder")


# =============================================================================
# Projetos (somente visualização)
# =============================================================================

def project_list(request):
    q = (request.GET.get("q") or "").strip()

    projects = Project.objects.all()
    try:
        projects = projects.exclude(cost_center__regex=r'^\s*0+\s*$')
    except Exception:
        projects = projects.exclude(cost_center__in=["0", "000000"])

    if q:
        projects = projects.filter(
            Q(name__icontains=q)
            | Q(code__icontains=q)
            | Q(cost_center__icontains=q)
            | Q(location__icontains=q)
        )

    page_obj = Paginator(projects, 12).get_page(request.GET.get("page"))
    return render(request, "engobras/projects/list.html", {"q": q, "page_obj": page_obj})


# =============================================================================
# EAP (somente visualização)
# =============================================================================

from django.db.models import Count, Prefetch, Q
from django.shortcuts import render
from .models import Milestone, GeneralActivity, SpecificActivity

def eap_view(request):
    specific_qs = (
        SpecificActivity.objects
        .filter(is_active=True)
        .order_by("order", "name")
    )

    general_qs = (
        GeneralActivity.objects
        .filter(is_active=True)
        .order_by("order", "name")
        .annotate(num_specifics=Count("engobras_specifics", filter=Q(engobras_specifics__is_active=True)))
        .prefetch_related(Prefetch("engobras_specifics", queryset=specific_qs))
    )

    milestones = (
        Milestone.objects
        .filter(is_active=True)
        .order_by("order", "name")
        .annotate(num_generals=Count("engobras_generals", filter=Q(engobras_generals__is_active=True)))
        .prefetch_related(Prefetch("engobras_generals", queryset=general_qs))
    )

    return render(request, "engobras/eap/index.html", {"milestones": milestones})




# =============================================================================
# Relatórios / Dashboards
# =============================================================================

@login_required
def reports_entry(request):
    """
    Entrada dos relatórios:
    - ?c=<id> -> redireciona direto para o dashboard do colaborador escolhido
    - se o usuário tiver Collaborator (ENGOBRAS) -> redireciona para o dashboard dele
    - caso contrário -> mostra lista de colaboradores para escolher
    """
    cid = request.GET.get("c")
    if cid and cid.isdigit():
        return redirect("engobras:collaborator_dashboard", pk=int(cid))

    me = _get_user_collaborator(request.user)
    if me:
        return redirect("engobras:collaborator_dashboard", pk=me.id)

    q = (request.GET.get("q") or "").strip()
    qs = Collaborator.objects.filter(is_active=True, department__slug=AUTO_DEPT_SLUG)
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q))

    page_obj = Paginator(qs.order_by("name"), 12).get_page(request.GET.get("page"))
    return render(request, "engobras/reports/select_collaborator.html", {
        "q": q,
        "page_obj": page_obj,
    })


@login_required
def collaborator_dashboard(request, pk: int):
    """Dashboard individual do colaborador com filtros e dados prontos para gráficos."""
    colab = get_object_or_404(Collaborator, pk=pk, department__slug=AUTO_DEPT_SLUG)

    # ------------------ Período (default: mês atual) ------------------
    today = timezone.localdate()
    default_start = today.replace(day=1)
    default_end = (default_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    date_from = _parse_date(request.GET.get("de") or "") or default_start
    date_to   = _parse_date(request.GET.get("ate") or "") or default_end
    if date_to < date_from:
        date_to = date_from

    period_start_dt = timezone.make_aware(datetime.combine(date_from, datetime.min.time()))
    period_end_dt   = timezone.make_aware(datetime.combine(date_to, datetime.max.time()))

    # ------------------ Filtros (multi-select) ------------------
    def _parse_ids(param):
        vals = request.GET.getlist(param)
        out = []
        for v in vals:
            try:
                out.append(int(v))
            except Exception:
                pass
        return out

    filt_projects   = _parse_ids("project")
    filt_milestones = _parse_ids("milestone")
    filt_generals   = _parse_ids("general")
    filt_specifics  = _parse_ids("specific")

    # Opções derivadas do que o colaborador já utilizou em Activity
    act_base = Activity.objects.filter(collaborator_id=colab.id)

    project_opts = (
        Project.objects
        .filter(id__in=act_base.values_list("project_id", flat=True))
        .exclude(id__isnull=True)
        .distinct().order_by("name")
        .values("id", "name")
    )

    milestone_opts = (
        Milestone.objects
        .filter(id__in=act_base.values_list("milestone_id", flat=True))
        .exclude(id__isnull=True)
        .distinct().order_by("order", "name")
        .values("id", "name")
    )

    general_opts = (
        GeneralActivity.objects
        .filter(id__in=act_base.values_list("general_id", flat=True))
        .exclude(id__isnull=True)
        .distinct().order_by("milestone__order", "order", "name")
        .values("id", "name")
    )

    specific_opts = (
        SpecificActivity.objects
        .filter(id__in=act_base.values_list("specific_id", flat=True))
        .exclude(id__isnull=True)
        .distinct().order_by("general__milestone__order", "general__order", "order", "name")
        .values("id", "name")
    )

    # ------------------ Query de sessões (com filtros) ------------------
    sessions = (
        ActivitySession.objects
        .filter(activity__collaborator_id=colab.id, started_at__lte=period_end_dt)
        .filter(Q(ended_at__isnull=True) | Q(ended_at__gte=period_start_dt))
        .select_related("activity", "activity__project", "activity__milestone",
                        "activity__general", "activity__specific")
        .order_by("started_at")
    )
    if filt_projects:
        sessions = sessions.filter(activity__project_id__in=filt_projects)
    if filt_milestones:
        sessions = sessions.filter(activity__milestone_id__in=filt_milestones)
    if filt_generals:
        sessions = sessions.filter(activity__general_id__in=filt_generals)
    if filt_specifics:
        sessions = sessions.filter(activity__specific_id__in=filt_specifics)

    # ------------------ Acúmulos ------------------
    total_seconds = 0
    per_day_secs = Counter()
    per_project_secs = defaultdict(int)
    per_milestone_secs = defaultdict(int)
    per_general_secs = defaultdict(int)
    per_specific_secs = defaultdict(int)

    # dias do range
    days = []
    d = date_from
    while d <= date_to:
        days.append(d)
        per_day_secs[d] = 0
        d += timedelta(days=1)

    for sess in sessions:
        sec = _overlap_seconds(sess.started_at, sess.ended_at, date_from, date_to)
        if sec <= 0:
            continue
        total_seconds += sec

        # Distribui por dia
        start = max(sess.started_at, period_start_dt)
        end = min(sess.ended_at or timezone.now(), period_end_dt)

        cursor_day = start.date()
        last_day = end.date()
        while cursor_day <= last_day:
            chunk = _overlap_seconds(start, end, cursor_day, cursor_day)
            if chunk > 0 and (date_from <= cursor_day <= date_to):
                per_day_secs[cursor_day] += chunk
            cursor_day += timedelta(days=1)

        # Chaves
        proj = sess.activity.project.name if sess.activity.project else "Sem projeto"
        per_project_secs[proj] += sec

        m = sess.activity.milestone.name if sess.activity.milestone_id else "—"
        g = sess.activity.general.name if sess.activity.general_id else "—"
        s = sess.activity.specific.name if sess.activity.specific_id else "—"

        per_milestone_secs[m] += sec
        per_general_secs[g] += sec
        per_specific_secs[s] += sec

    # ------------------ Tabelas + gráficos (dados prontos) ------------------
    def _sort_top(dct, limit):
        return sorted(dct.items(), key=lambda kv: kv[1], reverse=True)[:limit]

    # Horas por dia (labels/pcts)
    max_day = max(per_day_secs.values()) if per_day_secs else 0
    per_day_rows = []
    day_labels = []
    day_values = []  # em horas
    for day in days:
        secs = per_day_secs[day]
        pct = round(secs * 100 / max_day, 2) if max_day else 0.0
        per_day_rows.append({
            "date": day,
            "label": day.strftime("%d/%m"),
            "seconds": secs,
            "hms": _format_hms(secs),
            "pct": pct,
        })
        day_labels.append(day.strftime("%d/%m"))
        day_values.append(round(secs / 3600.0, 2))

    total_ref = total_seconds or 1
    top_projects = [{"name": n, "seconds": s, "hms": _format_hms(s),
                     "pct_total": round(s*100/total_ref, 2)}
                    for n, s in _sort_top(per_project_secs, 8)]
    top_eap_count = len(_sort_top(per_specific_secs, 10))
    distinct_projects_count = len(per_project_secs)

    proj_labels = [escape(n) for n, _ in _sort_top(per_project_secs, 8)]
    proj_hours  = [round(s / 3600.0, 2) for _, s in _sort_top(per_project_secs, 8)]

    mile_labels = [escape(n) for n, _ in _sort_top(per_milestone_secs, 10)]
    mile_hours  = [round(s / 3600.0, 2) for _, s in _sort_top(per_milestone_secs, 10)]

    gen_labels = [escape(n) for n, _ in _sort_top(per_general_secs, 10)]
    gen_hours  = [round(s / 3600.0, 2) for _, s in _sort_top(per_general_secs, 10)]

    chart_data = {
        "days": {"labels": day_labels, "values": day_values},
        "projects": {"labels": proj_labels, "values": proj_hours},
        "milestones": {"labels": mile_labels, "values": mile_hours},
        "generals": {"labels": gen_labels, "values": gen_hours},
    }

    context = {
        "colab": colab,
        "date_from": date_from,
        "date_to": date_to,
        "total_seconds": total_seconds,
        "total_hms": _format_hms(total_seconds),

        "project_opts": project_opts,
        "milestone_opts": milestone_opts,
        "general_opts": general_opts,
        "specific_opts": specific_opts,
        "filt_projects": set(filt_projects),
        "filt_milestones": set(filt_milestones),
        "filt_generals": set(filt_generals),
        "filt_specifics": set(filt_specifics),

        "per_day_rows": per_day_rows,
        "workdays_count": sum(1 for v in per_day_secs.values() if v > 0),
        "top_projects": top_projects,
        "distinct_projects_count": distinct_projects_count,
        "top_eap_count": top_eap_count,

        "chart_data": chart_data,
    }
    return render(request, "engobras/collaborators/dashboard.html", context)


# =============================================================================
# Colaboradores (somente visualização)
# =============================================================================

def collaborator_list(request):
    q = (request.GET.get("q") or "").strip()
    show_inactive = request.GET.get("inativos") == "1"

    qs = Collaborator.objects.filter(department__slug=AUTO_DEPT_SLUG)
    if not show_inactive:
        qs = qs.filter(is_active=True)

    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(email__icontains=q) |
            Q(phone__icontains=q)
        )

    page_obj = Paginator(qs.order_by("name"), 12).get_page(request.GET.get("page"))
    return render(request, "engobras/collaborators/list.html", {
        "q": q,
        "page_obj": page_obj,
        "show_inactive": show_inactive,
    })


@login_required
@require_POST
def close_workday(request):
    collab, err = _require_auto_collaborator(request)
    if err:
        return err

    today = timezone.localdate()
    wd = get_object_or_404(Workday, collaborator=collab, date=today)
    now = timezone.now()

    # 1) Encerra todas as sessões ainda abertas desse expediente
    ActivitySession.objects.filter(
        activity__workday=wd,
        ended_at__isnull=True
    ).update(ended_at=now)

    # 2) Marca as atividades do dia como INATIVAS
    Activity.objects.filter(
        workday=wd,
        finished_at__isnull=True,
        is_active=True
    ).update(is_active=False)

    # 3) Fecha o expediente
    wd.close()
    return redirect(f"{AUTO_NS}:dashboard")


# =============================================================================
# Configurações do Colaborador
# =============================================================================

@login_required
@transaction.atomic
def settings(request):
    collab, err = _require_auto_collaborator(request)
    if err:
        return err

    # forms padrão (GET)
    profile_form = CollaboratorSettingsForm(instance=collab)
    pwd_form = PasswordChangeForm(user=request.user)

    if request.method == "POST":
        # Descobre qual formulário foi submetido
        if "profile_submit" in request.POST:
            profile_form = CollaboratorSettingsForm(request.POST, instance=collab)
            if profile_form.is_valid():
                collab = profile_form.save()

                # Sincroniza o e-mail do User (opcional)
                if collab.user and collab.user.email != collab.email:
                    collab.user.email = collab.email
                    collab.user.save(update_fields=["email"])

                messages.success(request, "Suas configurações foram atualizadas com sucesso.")
                return redirect("engobras:settings")
            else:
                messages.error(request, "Revise os campos do formulário.")

        elif "password_submit" in request.POST:
            pwd_form = PasswordChangeForm(user=request.user, data=request.POST)
            if pwd_form.is_valid():
                user = pwd_form.save()
                update_session_auth_hash(request, user)  # mantém login ativo
                messages.success(request, "Sua senha foi alterada com sucesso.")
                return redirect("engobras:settings")
            else:
                messages.error(request, "Não foi possível alterar a senha. Verifique os campos.")

    return render(request, "engobras/configuration/settings.html", {
        "form": profile_form,
        "pwd_form": pwd_form,
    })


# =============================================================================
# Tempo real (somente gestores)
# =============================================================================

@login_required
def realtime(request):
    if not _user_is_manager(request.user):
        return HttpResponseForbidden("Acesso restrito a gestores.")
    activities = (Activity.objects
                  .select_related("collaborator", "project", "milestone", "general", "specific")
                  .filter(is_active=True, finished_at__isnull=True)
                  .order_by("collaborator__name", "-id"))
    return render(request, "engobras/realtime/index.html", {"activities": activities})


@login_required
def realtime_fragment(request):
    if not _user_is_manager(request.user):
        return HttpResponseForbidden("Acesso restrito a gestores.")
    activities = (Activity.objects
                  .select_related("collaborator", "project", "milestone", "general", "specific")
                  .filter(is_active=True, finished_at__isnull=True)
                  .order_by("collaborator__name", "-id"))
    return render(request, "engobras/realtime/_list.html", {"activities": activities})


# =============================================================================
# Relatório por departamento (gestores)
# =============================================================================

@login_required
def department_report(request):
    if not _user_is_manager(request.user):
        return HttpResponseForbidden("Acesso restrito a gestores.")

    # --- filtros básicos ---
    today = timezone.localdate()
    de = request.GET.get("de") or today.replace(day=1).isoformat()
    ate = request.GET.get("ate") or today.isoformat()

    # coletores de ids (múltiplos)
    def ids(param):
        vals = request.GET.getlist(param) or []
        return [int(v) for v in vals if v.isdigit()]

    filt_collabs   = ids("collaborator")
    filt_projects  = ids("project")
    filt_milestones= ids("milestone")
    filt_generals  = ids("general")
    filt_specifics = ids("specific")
    group_by = request.GET.get("group_by", "")
    order = request.GET.get("order", "-hours")

    # --- opções dos selects ---
    collaborator_opts = Collaborator.objects.filter(is_active=True, department__slug=AUTO_DEPT_SLUG).order_by("name")
    project_opts = Project.objects.order_by("name")
    milestone_opts = Milestone.objects.filter(is_active=True).order_by("order", "name")
    general_opts = GeneralActivity.objects.filter(is_active=True).order_by("milestone__order","order","name")
    specific_opts = SpecificActivity.objects.filter(is_active=True).order_by("general__milestone__order","general__order","order","name")

    # --- queryset base ---
    qs = (Activity.objects
          .select_related("collaborator","project","milestone","general","specific")
          .filter(started_at__date__gte=de, started_at__date__lte=ate))

    if filt_collabs:    qs = qs.filter(collaborator_id__in=filt_collabs)
    if filt_projects:   qs = qs.filter(project_id__in=filt_projects)
    if filt_milestones: qs = qs.filter(milestone_id__in=filt_milestones)
    if filt_generals:   qs = qs.filter(general_id__in=filt_generals)
    if filt_specifics:  qs = qs.filter(specific_id__in=filt_specifics)

    # --- agregações simples ---
    rows = []
    total_secs = 0

    per_day = defaultdict(int)
    per_collab = Counter()
    per_project = Counter()
    per_milestone = Counter()

    for a in qs:
        secs = a.total_active_seconds()
        total_secs += secs

        # dias (conta tudo no dia de início; para precisão, usar sessions/dia)
        day_key = a.started_at.date().isoformat()
        per_day[day_key] += secs

        # tops
        per_collab[a.collaborator.name] += secs
        per_project[(a.project.name if a.project else "Sem projeto")] += secs
        per_milestone[(a.milestone.name if a.milestone else "Sem marco")] += secs

        rows.append({
            "collaborator": a.collaborator.name,
            "project": a.project.name if a.project else None,
            "milestone": a.milestone.name if a.milestone else None,
            "general": a.general.name if a.general else None,
            "specific": a.specific.name if a.specific else None,
            "hours_h": f"{secs/3600:.2f} h",
            "hours": secs/3600.0,
        })

    # ordenar tabela
    if order == "-hours":
        rows.sort(key=lambda r: r["hours"], reverse=True)
    elif order == "hours":
        rows.sort(key=lambda r: r["hours"])
    elif order == "name":
        rows.sort(key=lambda r: r["collaborator"])
    elif order == "project":
        rows.sort(key=lambda r: (r["project"] or "ZZZ", r["collaborator"]))

    table_limit = 50
    table_rows = rows[:table_limit]

    # KPIs simples
    total_hms = f"{total_secs/3600:.2f} h"
    workdays_count = len({a.started_at.date() for a in qs})
    distinct_collaborators_count = len({a.collaborator_id for a in qs})
    distinct_projects_count = len({a.project_id for a in qs})

    # Chart data helpers
    def to_hours_list(counter_or_map):
        items = list(counter_or_map.items())
        items.sort(key=lambda kv: kv[1], reverse=True)
        labels = [k for k,_ in items[:10]]
        vals   = [round(v/3600,2) for _,v in items[:10]]
        return labels, vals

    # dias em ordem
    day_items = sorted(per_day.items(), key=lambda kv: kv[0])
    days_labels = [k for k,_ in day_items]
    days_vals   = [round(v/3600,2) for _,v in day_items]

    collab_labels, collab_vals = to_hours_list(per_collab)
    project_labels, project_vals = to_hours_list(per_project)
    milestone_labels, milestone_vals = to_hours_list(per_milestone)

    max_day_secs = max((v for _, v in day_items), default=0)
    per_day_rows = []
    for k, secs in day_items:  # k = 'YYYY-MM-DD'
        d = date.fromisoformat(k)
        per_day_rows.append({
            "label": d.strftime("%d/%m"),
            "seconds": secs,
            "hms": _format_hms(secs),
            "pct": round(secs * 100 / max_day_secs, 2) if max_day_secs else 0.0,
        })

    chart_data = {
        "days": {"labels": days_labels, "values": days_vals},
        "collaborators": {"labels": collab_labels, "values": collab_vals},
        "projects": {"labels": project_labels, "values": project_vals},
        "milestones": {"labels": milestone_labels, "values": milestone_vals},
    }

    ctx = {
        "date_from": date.fromisoformat(de),
        "date_to": date.fromisoformat(ate),

        "collaborator_opts": collaborator_opts,
        "project_opts": project_opts,
        "milestone_opts": milestone_opts,
        "general_opts": general_opts,
        "specific_opts": specific_opts,

        "filt_collaborators": filt_collabs,
        "filt_projects": filt_projects,
        "filt_milestones": filt_milestones,
        "filt_generals": filt_generals,
        "filt_specifics": filt_specifics,

        "group_by": group_by,
        "order": order,

        "total_hms": total_hms,
        "workdays_count": workdays_count,
        "distinct_collaborators_count": distinct_collaborators_count,
        "distinct_projects_count": distinct_projects_count,

        "table_rows": table_rows,
        "table_limit": table_limit,
        "chart_data": chart_data,
        "per_day_rows": per_day_rows,
    }
    return render(request, "engobras/reports/department_dashboard.html", ctx)


@login_required
def department_report_pdf(request):
    if not _user_is_manager(request.user):
        return HttpResponseForbidden("Acesso restrito a gestores.")
    # gere o PDF com os mesmos filtros de department_report ()
    # por enquanto, apenas um stub:
    return HttpResponse("PDF do Relatório Geral (ENGOBRAS) — implementar", content_type="text/plain")


# =============================================================================
# Exportações XLSX
# =============================================================================

@login_required
def collaborator_report_xlsx(request, pk):
    # 1) Leia os filtros (mesma lógica do dashboard)
    date_from = request.GET.get("de")
    date_to   = request.GET.get("ate")
    projects  = request.GET.getlist("project")
    miles     = request.GET.getlist("milestone")
    generals  = request.GET.getlist("general")
    specifics = request.GET.getlist("specific")

    # (permissões) gestor ou dono do relatório
    me = _get_user_collaborator(request.user)
    if me and (not _user_is_manager(request.user)) and me.id != int(pk):
        return HttpResponseForbidden("Você não tem permissão para exportar este colaborador.")

    # 2) Query base
    qs = (Activity.objects
          .select_related("project","milestone","general","specific","workday","collaborator")
          .filter(collaborator_id=pk))
    if date_from:
        qs = qs.filter(Q(workday__date__gte=date_from) | Q(workday__isnull=True, started_at__date__gte=date_from))
    if date_to:
        qs = qs.filter(Q(workday__date__lte=date_to) | Q(workday__isnull=True, started_at__date__lte=date_to))
    if projects:  qs = qs.filter(project_id__in=projects)
    if miles:     qs = qs.filter(milestone_id__in=miles)
    if generals:  qs = qs.filter(general_id__in=generals)
    if specifics: qs = qs.filter(specific_id__in=specifics)

    # 3) Monte o XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    headers = [
        "Data", "Projeto", "Marco", "Geral", "Específica",
        "Descrição", "Horas (dec)", "Horas (HH:MM:SS)"
    ]
    ws.append(headers)

    for a in qs:
        secs = a.total_active_seconds()
        h = secs/3600
        hh = f"{secs//3600:02d}:{(secs%3600)//60:02d}:{secs%60:02d}"
        ws.append([
            a.workday.date.strftime("%Y-%m-%d") if a.workday_id else a.started_at.date().isoformat(),
            a.project.name if a.project else "",
            a.milestone.name if a.milestone else (getattr(a, "custom_milestone", "") or ""),
            a.general.name if a.general else (getattr(a, "custom_general", "") or ""),
            a.specific.name if a.specific else (getattr(a, "custom_specific", "") or ""),
            a.description or "",
            round(h, 2),
            hh,
        ])

    # 4) Larguras auto-ajustadas simples
    for col in range(1, ws.max_column+1):
        length = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row+1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, length + 2), 60)

    # 5) Resposta
    filename = f"relatorio_colaborador_{pk}_{dt.date.today().isoformat()}.xlsx"
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(resp)
    return resp


@login_required
def global_report_xlsx(request):
    # Mesma leitura de filtros do department_report
    today = timezone.localdate()
    de = request.GET.get("de") or today.replace(day=1).isoformat()
    ate = request.GET.get("ate") or today.isoformat()

    def ids(param):
        vals = request.GET.getlist(param) or []
        return [int(v) for v in vals if v.isdigit()]

    filt_collabs   = ids("collaborator")
    filt_projects  = ids("project")
    filt_milestones= ids("milestone")
    filt_generals  = ids("general")
    filt_specifics = ids("specific")

    qs = (Activity.objects
          .select_related("collaborator","project","milestone","general","specific","workday")
          .filter(started_at__date__gte=de, started_at__date__lte=ate))

    if filt_collabs:    qs = qs.filter(collaborator_id__in=filt_collabs)
    if filt_projects:   qs = qs.filter(project_id__in=filt_projects)
    if filt_milestones: qs = qs.filter(milestone_id__in=filt_milestones)
    if filt_generals:   qs = qs.filter(general_id__in=filt_generals)
    if filt_specifics:  qs = qs.filter(specific_id__in=filt_specifics)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Geral"

    headers = [
        "Data", "Colaborador", "Projeto", "Marco", "Geral", "Específica",
        "Descrição", "Horas (dec)", "Horas (HH:MM:SS)"
    ]
    ws.append(headers)

    for a in qs:
        secs = a.total_active_seconds()
        ws.append([
            (a.workday.date.strftime("%Y-%m-%d") if a.workday_id else a.started_at.date().isoformat()),
            a.collaborator.name if a.collaborator_id else "",
            a.project.name if a.project else "",
            a.milestone.name if a.milestone else (getattr(a, "custom_milestone", "") or ""),
            a.general.name if a.general else (getattr(a, "custom_general", "") or ""),
            a.specific.name if a.specific else (getattr(a, "custom_specific", "") or ""),
            a.description or "",
            round(secs/3600, 2),
            f"{secs//3600:02d}:{(secs%3600)//60:02d}:{secs%60:02d}",
        ])

    # Larguras
    for col in range(1, ws.max_column+1):
        length = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row+1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, length + 2), 60)

    filename = f"relatorio_geral_{dt.date.today().isoformat()}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(resp)
    return resp


@login_required
def catalogs(request):
    generals = GeneralActivity.objects.filter(is_active=True).order_by("order", "name")
    panel_sizes = PanelSize.objects.filter(is_active=True).order_by("order", "name")

    ctx = {
        "generals": generals,
        "panel_sizes": panel_sizes,
    }
    return render(request, "engobras/epe/catalogs/index.html", ctx)
