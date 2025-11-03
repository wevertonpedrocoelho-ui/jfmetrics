# epe/views.py

from datetime import datetime, timedelta, date
from collections import defaultdict, Counter
import datetime as dt

from django.apps import apps
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.html import escape
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# >>> usa os forms do EPE (ActivityForm) e o form de perfil apontando para common.Collaborator
from .forms import ActivityForm, CollaboratorSettingsForm

# >>> modelos do próprio EPE
from .models import (
    Activity,
    ActivitySession,
    GeneralActivity,
    PanelSize,
    Project,
    Workday,
)

# >>> colaborador centralizado
from common.models import Collaborator, Department

# =============================================================================
# Constantes / helpers
# =============================================================================

DEPARTMENT_NAME = apps.get_app_config("epe").verbose_name  # "Engenharia de Painéis Elétricos"
EPE_DEPT_SLUG = "epe"  # ajuste se o slug do seu Departamento for outro


def _get_user_collaborator(user):
    """
    Retorna o Collaborator central do usuário para o departamento EPE.
    NÃO cria placeholder. Se não existir, retorna None.
    """
    if not user or not getattr(user, "is_authenticated", False):
        return None
    return (
        Collaborator.objects
        .select_related("department", "user")
        .filter(user=user, is_active=True, department__slug=EPE_DEPT_SLUG)
        .first()
    )


def _require_epe_collaborator(request):
    """
    Busca o colaborador EPE do usuário atual.
    Se não existir, nega acesso (403). Use onde a presença é obrigatória.
    """
    collab = _get_user_collaborator(request.user)
    if not collab:
        return None, HttpResponseForbidden("Você não possui perfil neste departamento.")
    return collab, None


def get_user_activity_or_404(user, pk: int):
    """Garante que a Activity pertence ao colaborador EPE do usuário atual."""
    collab = _get_user_collaborator(user)
    return get_object_or_404(Activity, pk=pk, collaborator=collab)


def _parse_date(s: str):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def _overlap_seconds(start_a, end_a, start_b_date, end_b_date):
    """
    Segundos de sobreposição entre [start_a, end_a] e o dia (ou intervalo de dias)
    [start_b_date 00:00:00, end_b_date 23:59:59].
    """
    if end_a is None:
        end_a = timezone.now()
    start_b = timezone.make_aware(datetime.combine(start_b_date, datetime.min.time()))
    end_b = timezone.make_aware(datetime.combine(end_b_date, datetime.max.time()))
    start = max(start_a, start_b)
    end = min(end_a, end_b)
    delta = (end - start).total_seconds()
    return int(delta) if delta > 0 else 0


def _format_hms(total_seconds: int) -> str:
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def _user_is_manager(user) -> bool:
    try:
        c = _get_user_collaborator(user)
        return bool(c and c.is_manager)
    except Exception:
        return False


# =============================================================================
# Dashboard / expediente
# =============================================================================

@login_required
def dashboard(request):
    collab = _get_user_collaborator(request.user)

    # Data selecionada via query (?date=YYYY-MM-DD)
    today = timezone.localdate()
    date_str = request.GET.get("date", "")
    try:
        selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else today
    except ValueError:
        selected_date = today
    is_today = selected_date == today

    workday = None
    activities = Activity.objects.none()
    stats = {"total_secs": 0, "active": 0, "paused": 0, "done": 0}

    if collab:
        workday = Workday.objects.filter(collaborator=collab, date=selected_date).first()
        if workday:
            activities = (
                Activity.objects.filter(collaborator=collab, workday=workday)
                .select_related("project", "general", "panel_size")
                .prefetch_related("sessions")
                .order_by("-id")
            )
            stats["total_secs"] = sum(a.total_active_seconds() for a in activities)
            stats["active"] = activities.filter(is_active=True, finished_at__isnull=True).count()
            stats["paused"] = activities.filter(is_active=False, finished_at__isnull=True).count()
            stats["done"] = activities.filter(finished_at__isnull=False).count()

    # prev/next (não permite passar de hoje)
    prev_date = selected_date - timedelta(days=1)
    next_date = selected_date + timedelta(days=1)
    can_next = next_date <= today

    # faixa de dias (pílulas)
    window = []
    span = range(-3, 4)  # 7 dias
    existing_dates = set()
    if collab:
        dates_in_window = [
            selected_date + timedelta(days=i) for i in span if (selected_date + timedelta(days=i)) <= today
        ]
        existing_dates = set(
            Workday.objects.filter(collaborator=collab, date__in=dates_in_window).values_list("date", flat=True)
        )
    for i in span:
        d = selected_date + timedelta(days=i)
        if d > today:
            continue
        window.append(
            {
                "date": d,
                "iso": d.isoformat(),
                "label": d.strftime("%d/%m"),
                "is_selected": d == selected_date,
                "is_today": d == today,
                "has_workday": d in existing_dates,
            }
        )

    context = {
        "department_name": DEPARTMENT_NAME,
        "collab": collab,
        "workday": workday,
        "activities": activities,
        "stats": stats,
        "projects": Project.objects.order_by("name"),
        "selected_date": selected_date,
        "is_today": is_today,
        "prev_date": prev_date,
        "next_date": next_date if can_next else None,
        "nav_days": window,

        "generals": GeneralActivity.objects.filter(is_active=True).order_by("order", "name"),
        "panel_sizes": PanelSize.objects.filter(is_active=True).order_by("order", "name"),
    }
    return render(request, "epe/dashboard.html", context)


@login_required
@require_POST
def start_workday(request):
    collab, err = _require_epe_collaborator(request)
    if err:
        return err
    wd, created = Workday.objects.get_or_create(
        collaborator=collab,
        date=timezone.localdate(),
        defaults={"is_open": True},
    )
    if not created and not wd.is_open:
        wd.is_open, wd.ended_at = True, None
        wd.save(update_fields=["is_open", "ended_at"])
    return redirect("epe:dashboard")


# =============================================================================
# Atividades
# =============================================================================

@login_required
@transaction.atomic
def activity_create(request):
    """
    Cria a atividade e já inicia a primeira sessão (cronômetro).
    O ActivityForm do app epe deve ter fields: project, panel_name, general, panel_size, description.
    """
    collab, err = _require_epe_collaborator(request)
    if err:
        return err

    wd, _ = Workday.objects.get_or_create(
        collaborator=collab,
        date=timezone.localdate(),
        defaults={"is_open": True},
    )

    form = ActivityForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        act: Activity = form.save(commit=False)
        act.collaborator = collab
        act.workday = wd
        act.save()
        # inicia cronômetro
        ActivitySession.objects.create(activity=act)
        return redirect("epe:dashboard")

    # (corrigido o path do template)
    return render(request, "epe/epe/activity_form.html", {"form": form})


@login_required
@require_POST
def activity_start(request, pk):
    act = get_user_activity_or_404(request.user, pk)
    last = act.sessions.order_by("-id").first()
    if not last or last.ended_at:
        ActivitySession.objects.create(activity=act)
    act.is_active = True
    act.save(update_fields=["is_active"])
    return redirect("epe:dashboard")


@login_required
@require_POST
def activity_pause(request, pk):
    act = get_user_activity_or_404(request.user, pk)
    last = act.sessions.order_by("-id").first()
    if last and not last.ended_at:
        last.stop()
    act.is_active = False
    act.save(update_fields=["is_active"])
    return redirect("epe:dashboard")


@login_required
@require_POST
def activity_resume(request, pk):
    return activity_start(request, pk)


@login_required
@require_POST
def activity_finish(request, pk):
    act = get_user_activity_or_404(request.user, pk)
    for s in act.sessions.filter(ended_at__isnull=True):
        s.stop()
    act.is_active = False
    act.finished_at = timezone.now()
    act.save(update_fields=["is_active", "finished_at"])
    return redirect("epe:dashboard")


# =============================================================================
# Páginas simples
# =============================================================================

def profile(request):
    return HttpResponse("Perfil (EPE) — placeholder")


# =============================================================================
# Projetos (somente visualização)
# =============================================================================

def project_list(request):
    q = (request.GET.get("q") or "").strip()

    projects = Project.objects.all()
    try:
        projects = projects.exclude(cost_center__regex=r"^\s*0+\s*$")
    except Exception:
        projects = projects.exclude(cost_center__in=["0", "000000"])

    if q:
        projects = projects.filter(
            Q(name__icontains=q)
            | Q(code__icontains=q)
            | Q(cost_center__icontains=q)
            | Q(location__icontains=q)
        )

    page_obj = Paginator(projects, 12).get_page(request.GET.get("page"))
    return render(request, "epe/projects/list.html", {"q": q, "page_obj": page_obj})


# =============================================================================
# Relatórios / Dashboards
# =============================================================================

@login_required
def reports_entry(request):
    """
    Entrada dos relatórios do app EPE:
    - ?c=<id> -> redireciona direto para o dashboard do colaborador escolhido
    - se o usuário tiver Collaborator (EPE) vinculado -> redireciona para o dashboard dele
    - caso contrário -> mostra lista de colaboradores (EPE) para escolher
    """
    cid = request.GET.get("c")
    if cid and cid.isdigit():
        return redirect("epe:collaborator_dashboard", pk=int(cid))

    me = _get_user_collaborator(request.user)
    if me:
        return redirect("epe:collaborator_dashboard", pk=me.id)

    q = (request.GET.get("q") or "").strip()
    qs = Collaborator.objects.filter(is_active=True, department__slug=EPE_DEPT_SLUG)

    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q))

    page_obj = Paginator(qs.order_by("name"), 12).get_page(request.GET.get("page"))
    return render(
        request,
        "epe/reports/select_collaborator.html",
        {
            "q": q,
            "page_obj": page_obj,
            "department_name": DEPARTMENT_NAME,
        },
    )


@login_required
@login_required
def collaborator_dashboard(request, pk: int):
    """Dashboard individual do colaborador (EPE) com filtros e dados prontos para gráficos."""
    colab = get_object_or_404(Collaborator, pk=pk, department__slug=EPE_DEPT_SLUG)

    # ------------------ Período (default: mês atual) ------------------
    today = timezone.localdate()
    default_start = today.replace(day=1)
    default_end = (default_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    date_from = _parse_date(request.GET.get("de") or "") or default_start
    date_to = _parse_date(request.GET.get("ate") or "") or default_end
    if date_to < date_from:
        date_to = date_from

    period_start_dt = timezone.make_aware(datetime.combine(date_from, datetime.min.time()))
    period_end_dt = timezone.make_aware(datetime.combine(date_to, datetime.max.time()))

    # ------------------ Filtros (multi-select) ------------------
    def _parse_ids(param):
        vals = request.GET.getlist(param)
        out = []
        for v in vals:
            try:
                out.append(int(v))
            except Exception:
                pass
        return out

    filt_projects = _parse_ids("project")
    filt_generals = _parse_ids("general")
    filt_sizes    = _parse_ids("size")
    # nomes (CharField) — não são IDs
    filt_panels   = [v for v in request.GET.getlist("panel") if v.strip()]

    # Opções derivadas do que o colaborador já utilizou em Activity
    act_base = Activity.objects.filter(collaborator_id=colab.id)

    project_opts = (
        Project.objects.filter(id__in=act_base.values_list("project_id", flat=True))
        .exclude(id__isnull=True)
        .distinct().order_by("name").values("id", "name")
    )
    general_opts = (
        GeneralActivity.objects.filter(id__in=act_base.values_list("general_id", flat=True))
        .exclude(id__isnull=True)
        .distinct().order_by("order", "name").values("id", "name")
    )
    size_opts = (
        PanelSize.objects.filter(id__in=act_base.values_list("panel_size_id", flat=True))
        .exclude(id__isnull=True)
        .distinct().order_by("order", "name").values("id", "name")
    )
    # novos: nomes de painel distintos já usados pelo colaborador
    panel_name_opts = (
        act_base.values_list("panel_name", flat=True)
        .distinct().order_by("panel_name")
    )

    # ------------------ Query de sessões (com filtros) ------------------
    sessions = (
        ActivitySession.objects
        .filter(activity__collaborator_id=colab.id, started_at__lte=period_end_dt)
        .filter(Q(ended_at__isnull=True) | Q(ended_at__gte=period_start_dt))
        .select_related("activity", "activity__project", "activity__general", "activity__panel_size")
        .order_by("started_at")
    )
    if filt_projects:
        sessions = sessions.filter(activity__project_id__in=filt_projects)
    if filt_generals:
        sessions = sessions.filter(activity__general_id__in=filt_generals)
    if filt_sizes:
        sessions = sessions.filter(activity__panel_size_id__in=filt_sizes)
    if filt_panels:
        sessions = sessions.filter(activity__panel_name__in=filt_panels)

    # ------------------ Acúmulos ------------------
    total_seconds = 0
    per_day_secs = Counter()
    per_project_secs = defaultdict(int)
    per_general_secs = defaultdict(int)
    per_size_secs = defaultdict(int)
    per_panel_secs = defaultdict(int)  # <-- novo

    # dias do range
    days = []
    d = date_from
    while d <= date_to:
        days.append(d)
        per_day_secs[d] = 0
        d += timedelta(days=1)

    for sess in sessions:
        sec = _overlap_seconds(sess.started_at, sess.ended_at, date_from, date_to)
        if sec <= 0:
            continue
        total_seconds += sec

        # Distribui por dia
        start = max(sess.started_at, period_start_dt)
        end = min(sess.ended_at or timezone.now(), period_end_dt)

        cursor_day = start.date()
        last_day = end.date()
        while cursor_day <= last_day:
            chunk = _overlap_seconds(start, end, cursor_day, cursor_day)
            if chunk > 0 and (date_from <= cursor_day <= date_to):
                per_day_secs[cursor_day] += chunk
            cursor_day += timedelta(days=1)

        # Chaves
        proj = sess.activity.project.name if sess.activity.project else "Sem projeto"
        per_project_secs[proj] += sec

        g = sess.activity.general.name if sess.activity.general_id else "—"
        s = sess.activity.panel_size.name if sess.activity.panel_size_id else "—"
        p = sess.activity.panel_name or "—"

        per_general_secs[g] += sec
        per_size_secs[s] += sec
        per_panel_secs[p] += sec  # <-- novo

    # ------------------ Tabelas + gráficos (dados prontos) ------------------
    def _sort_top(dct, limit):
        return sorted(dct.items(), key=lambda kv: kv[1], reverse=True)[:limit]

    # Horas por dia (labels/pcts)
    max_day = max(per_day_secs.values()) if per_day_secs else 0
    per_day_rows = []
    day_labels = []
    day_values = []  # em horas
    for day in days:
        secs = per_day_secs[day]
        pct = round(secs * 100 / max_day, 2) if max_day else 0.0
        per_day_rows.append(
            {"date": day, "label": day.strftime("%d/%m"), "seconds": secs, "hms": _format_hms(secs), "pct": pct}
        )
        day_labels.append(day.strftime("%d/%m"))
        day_values.append(round(secs / 3600.0, 2))

    total_ref = total_seconds or 1
    top_projects = [
        {"name": n, "seconds": s, "hms": _format_hms(s), "pct_total": round(s * 100 / total_ref, 2)}
        for n, s in _sort_top(per_project_secs, 8)
    ]
    distinct_projects_count = len(per_project_secs)

    proj_labels = [escape(n) for n, _ in _sort_top(per_project_secs, 8)]
    proj_hours = [round(s / 3600.0, 2) for _, s in _sort_top(per_project_secs, 8)]

    gen_labels = [escape(n) for n, _ in _sort_top(per_general_secs, 10)]
    gen_hours = [round(s / 3600.0, 2) for _, s in _sort_top(per_general_secs, 10)]

    size_labels = [escape(n) for n, _ in _sort_top(per_size_secs, 10)]
    size_hours = [round(s / 3600.0, 2) for _, s in _sort_top(per_size_secs, 10)]

    # Top painéis (novo)
    panel_labels = [escape(n) for n, _ in _sort_top(per_panel_secs, 10)]
    panel_hours = [round(s / 3600.0, 2) for _, s in _sort_top(per_panel_secs, 10)]
    distinct_panels_count = len(per_panel_secs)

    chart_data = {
        "days": {"labels": day_labels, "values": day_values},
        "projects": {"labels": proj_labels, "values": proj_hours},
        "generals": {"labels": gen_labels, "values": gen_hours},
        "sizes": {"labels": size_labels, "values": size_hours},
        "panels": {"labels": panel_labels, "values": panel_hours},  # <-- novo
    }

    context = {
        "department_name": DEPARTMENT_NAME,
        "colab": colab,
        "date_from": date_from,
        "date_to": date_to,
        "total_seconds": total_seconds,
        "total_hms": _format_hms(total_seconds),

        "project_opts": project_opts,
        "general_opts": general_opts,
        "size_opts": size_opts,
        "panel_name_opts": panel_name_opts,          # <-- novo

        "filt_projects": set(filt_projects),
        "filt_generals": set(filt_generals),
        "filt_sizes": set(filt_sizes),
        "filt_panels": set(filt_panels),            # <-- novo

        "per_day_rows": per_day_rows,
        "workdays_count": sum(1 for v in per_day_secs.values() if v > 0),
        "top_projects": top_projects,
        "distinct_projects_count": distinct_projects_count,
        "distinct_panels_count": distinct_panels_count,  # <-- novo
        "chart_data": chart_data,
    }
    # mantém o caminho atual do seu projeto:
    return render(request, "epe/collaborators/dashboard.html", context)



# =============================================================================
# Colaboradores (somente visualização)
# =============================================================================

def collaborator_list(request):
    q = (request.GET.get("q") or "").strip()
    show_inactive = request.GET.get("inativos") == "1"

    qs = Collaborator.objects.filter(department__slug=EPE_DEPT_SLUG)
    if not show_inactive:
        qs = qs.filter(is_active=True)

    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q))

    page_obj = Paginator(qs.order_by("name"), 12).get_page(request.GET.get("page"))
    return render(
        request,
        "epe/collaborators/list.html",
        {
            "q": q,
            "page_obj": page_obj,
            "show_inactive": show_inactive,
            "department_name": DEPARTMENT_NAME,
        },
    )


@login_required
@require_POST
def close_workday(request):
    collab, err = _require_epe_collaborator(request)
    if err:
        return err

    today = timezone.localdate()
    wd = get_object_or_404(Workday, collaborator=collab, date=today)
    now = timezone.now()

    # 1) Encerra todas as sessões ainda abertas desse expediente
    ActivitySession.objects.filter(activity__workday=wd, ended_at__isnull=True).update(ended_at=now)

    # 2) Marca as atividades do dia como INATIVAS
    Activity.objects.filter(workday=wd, finished_at__isnull=True, is_active=True).update(is_active=False)

    # 3) Fecha o expediente
    wd.close()
    return redirect("epe:dashboard")


# =============================================================================
# Configurações do Colaborador (perfil + troca de senha)
# =============================================================================

@login_required
@transaction.atomic
def settings(request):
    collab, err = _require_epe_collaborator(request)
    if err:
        return err

    # forms default (GET)
    profile_form = CollaboratorSettingsForm(instance=collab)
    pwd_form = PasswordChangeForm(user=request.user)

    if request.method == "POST":
        if "profile_submit" in request.POST:
            profile_form = CollaboratorSettingsForm(request.POST, instance=collab)
            if profile_form.is_valid():
                collab = profile_form.save()
                # Sincroniza e-mail do User (opcional)
                if collab.user and collab.user.email != collab.email:
                    collab.user.email = collab.email
                    collab.user.save(update_fields=["email"])
                messages.success(request, "Suas configurações foram atualizadas com sucesso.")
                return redirect("epe:settings")
            else:
                messages.error(request, "Revise os campos do formulário.")
        elif "password_submit" in request.POST:
            pwd_form = PasswordChangeForm(user=request.user, data=request.POST)
            if pwd_form.is_valid():
                user = pwd_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Sua senha foi alterada com sucesso.")
                return redirect("epe:settings")
            else:
                messages.error(request, "Não foi possível alterar a senha. Verifique os campos.")

    return render(
        request,
        "epe/configuration/settings.html",
        {
            "form": profile_form,
            "pwd_form": pwd_form,
            "department_name": DEPARTMENT_NAME,
        },
    )


# =============================================================================
# Tempo real (somente gestores)
# =============================================================================

@login_required
def realtime(request):
    if not _user_is_manager(request.user):
        return HttpResponseForbidden("Acesso restrito a gestores.")
    activities = (
        Activity.objects.select_related("collaborator", "project", "general", "panel_size")
        .filter(is_active=True, finished_at__isnull=True)
        .order_by("collaborator__name", "-id")
    )
    return render(
        request,
        "epe/realtime/index.html",
        {"activities": activities, "department_name": DEPARTMENT_NAME},
    )


@login_required
def realtime_fragment(request):
    if not _user_is_manager(request.user):
        return HttpResponseForbidden("Acesso restrito a gestores.")
    activities = (
        Activity.objects.select_related("collaborator", "project", "general", "panel_size")
        .filter(is_active=True, finished_at__isnull=True)
        .order_by("collaborator__name", "-id")
    )
    return render(request, "epe/realtime/_list.html", {"activities": activities})


# =============================================================================
# Relatório por departamento (gestores)
# =============================================================================

@login_required
def department_report(request):
    if not _user_is_manager(request.user):
        return HttpResponseForbidden("Acesso restrito a gestores.")

    # --- filtros básicos ---
    today = timezone.localdate()
    de = request.GET.get("de") or today.replace(day=1).isoformat()
    ate = request.GET.get("ate") or today.isoformat()

    # coletores de ids (múltiplos)
    def ids(param):
        vals = request.GET.getlist(param) or []
        return [int(v) for v in vals if v.isdigit()]

    filt_collabs = ids("collaborator")
    filt_projects = ids("project")
    filt_generals = ids("general")
    filt_sizes = ids("size")
    group_by = request.GET.get("group_by", "")
    order = request.GET.get("order", "-hours")

    # --- opções dos selects ---
    collaborator_opts = Collaborator.objects.filter(is_active=True, department__slug=EPE_DEPT_SLUG).order_by("name")
    project_opts = Project.objects.order_by("name")
    general_opts = GeneralActivity.objects.filter(is_active=True).order_by("order", "name")
    size_opts = PanelSize.objects.filter(is_active=True).order_by("order", "name")

    # --- queryset base ---
    qs = (
        Activity.objects.select_related("collaborator", "project", "general", "panel_size").filter(
            started_at__date__gte=de, started_at__date__lte=ate
        )
    )

    if filt_collabs:
        qs = qs.filter(collaborator_id__in=filt_collabs)
    if filt_projects:
        qs = qs.filter(project_id__in=filt_projects)
    if filt_generals:
        qs = qs.filter(general_id__in=filt_generals)
    if filt_sizes:
        qs = qs.filter(panel_size_id__in=filt_sizes)

    # --- agregações simples ---
    rows = []
    total_secs = 0

    per_day = defaultdict(int)
    per_collab = Counter()
    per_project = Counter()
    per_general = Counter()
    per_size = Counter()

    for a in qs:
        secs = a.total_active_seconds()
        total_secs += secs

        day_key = a.started_at.date().isoformat()
        per_day[day_key] += secs

        # tops
        per_collab[a.collaborator.name] += secs
        per_project[(a.project.name if a.project else "Sem projeto")] += secs
        per_general[(a.general.name if a.general else "—")] += secs
        per_size[(a.panel_size.name if a.panel_size else "—")] += secs

        rows.append(
            {
                "collaborator": a.collaborator.name,
                "project": a.project.name if a.project else None,
                "general": a.general.name if a.general else None,
                "panel_size": a.panel_size.name if a.panel_size else None,
                "panel_name": a.panel_name,
                "hours_h": f"{secs/3600:.2f} h",
                "hours": secs / 3600.0,
            }
        )

    # ordenar tabela
    if order == "-hours":
        rows.sort(key=lambda r: r["hours"], reverse=True)
    elif order == "hours":
        rows.sort(key=lambda r: r["hours"])
    elif order == "name":
        rows.sort(key=lambda r: r["collaborator"])
    elif order == "project":
        rows.sort(key=lambda r: (r["project"] or "ZZZ", r["collaborator"]))

    table_limit = 50
    table_rows = rows[:table_limit]

    # KPIs simples
    total_hms = f"{total_secs/3600:.2f} h"
    workdays_count = len({a.started_at.date() for a in qs})
    distinct_collaborators_count = len({a.collaborator_id for a in qs})
    distinct_projects_count = len({a.project_id for a in qs})

    # Chart data helpers
    def to_hours_list(counter_or_map):
        items = list(counter_or_map.items())
        items.sort(key=lambda kv: kv[1], reverse=True)
        labels = [k for k, _ in items[:10]]
        vals = [round(v / 3600, 2) for _, v in items[:10]]
        return labels, vals

    # dias em ordem
    day_items = sorted(per_day.items(), key=lambda kv: kv[0])
    days_labels = [k for k, _ in day_items]
    days_vals = [round(v / 3600, 2) for _, v in day_items]

    collab_labels, collab_vals = to_hours_list(per_collab)
    project_labels, project_vals = to_hours_list(per_project)
    general_labels, general_vals = to_hours_list(per_general)
    size_labels, size_vals = to_hours_list(per_size)

    max_day_secs = max((v for _, v in day_items), default=0)
    per_day_rows = []
    for k, secs in day_items:  # k = 'YYYY-MM-DD'
        d = date.fromisoformat(k)
        per_day_rows.append(
            {
                "label": d.strftime("%d/%m"),
                "seconds": secs,
                "hms": _format_hms(secs),
                "pct": round(secs * 100 / max_day_secs, 2) if max_day_secs else 0.0,
            }
        )

    chart_data = {
        "days": {"labels": days_labels, "values": days_vals},
        "collaborators": {"labels": collab_labels, "values": collab_vals},
        "projects": {"labels": project_labels, "values": project_vals},
        "generals": {"labels": general_labels, "values": general_vals},
        "sizes": {"labels": size_labels, "values": size_vals},
    }

    ctx = {
        "department_name": DEPARTMENT_NAME,
        "date_from": date.fromisoformat(de),
        "date_to": date.fromisoformat(ate),
        "collaborator_opts": collaborator_opts,
        "project_opts": project_opts,
        "general_opts": general_opts,
        "size_opts": size_opts,
        "filt_collaborators": filt_collabs,
        "filt_projects": filt_projects,
        "filt_generals": filt_generals,
        "filt_sizes": filt_sizes,
        "group_by": group_by,
        "order": order,
        "total_hms": total_hms,
        "workdays_count": workdays_count,
        "distinct_collaborators_count": distinct_collaborators_count,
        "distinct_projects_count": distinct_projects_count,
        "table_rows": table_rows,
        "table_limit": table_limit,
        "chart_data": chart_data,
        "per_day_rows": per_day_rows,
    }
    return render(request, "epe/reports/department_dashboard.html", ctx)


@login_required
def department_report_pdf(request):
    if not _user_is_manager(request.user):
        return HttpResponseForbidden("Acesso restrito a gestores.")
    # gere o PDF com os mesmos filtros de department_report ()
    # por enquanto, apenas um stub:
    return HttpResponse("PDF do Relatório Geral (EPE) — implementar", content_type="text/plain")


# =============================================================================
# Exportações XLSX (EPE)
# =============================================================================

@login_required
def collaborator_report_xlsx(request, pk: int):
    """
    Exporta as atividades do colaborador (EPE) no período e filtros informados.
    Colunas: Data, Projeto, Atividade Geral, Tamanho do painel, Nome do painel, Descrição, Horas (dec), Horas (HH:MM:SS)
    """
    # Filtros (mesma ideia do dashboard)
    date_from = request.GET.get("de")
    date_to = request.GET.get("ate")
    projects = request.GET.getlist("project")
    generals = request.GET.getlist("general")
    sizes = request.GET.getlist("size")
    panels = [v for v in request.GET.getlist("panel") if v.strip()]  # <-- novo

    # Permissão: gestor OU o próprio colaborador
    me = _get_user_collaborator(request.user)
    if me and (not _user_is_manager(request.user)) and me.id != int(pk):
        return HttpResponseForbidden("Você não tem permissão para exportar este colaborador.")

    # Query base
    qs = (
        Activity.objects.select_related("project", "general", "panel_size", "workday", "collaborator")
        .filter(collaborator_id=pk)
        .order_by("started_at", "id")
    )
    if date_from:
        qs = qs.filter(Q(workday__date__gte=date_from) | Q(workday__isnull=True, started_at__date__gte=date_from))
    if date_to:
        qs = qs.filter(Q(workday__date__lte=date_to) | Q(workday__isnull=True, started_at__date__lte=date_to))
    if projects:
        qs = qs.filter(project_id__in=projects)
    if generals:
        qs = qs.filter(general_id__in=generals)
    if sizes:
        qs = qs.filter(panel_size_id__in=sizes)
    if panels:
        qs = qs.filter(panel_name__in=panels)  # <-- novo

    # Monta XLSX (restante preservado) ...
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    headers = [
        "Data",
        "Projeto",
        "Atividade Geral",
        "Tamanho do painel",
        "Nome do painel",
        "Descrição",
        "Horas (dec)",
        "Horas (HH:MM:SS)",
    ]
    ws.append(headers)

    for a in qs:
        secs = a.total_active_seconds()
        hours_dec = round(secs / 3600.0, 2)
        hh = f"{secs//3600:02d}:{(secs%3600)//60:02d}:{secs%60:02d}"

        date_str = a.workday.date.strftime("%Y-%m-%d") if getattr(a, "workday_id", None) else a.started_at.date().isoformat()

        ws.append(
            [
                date_str,
                (a.project.name if a.project else ""),
                (a.general.name if a.general_id else ""),
                (a.panel_size.name if a.panel_size_id else ""),
                (a.panel_name or ""),
                (a.description or ""),
                hours_dec,
                hh,
            ]
        )

    for col in range(1, ws.max_column + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 60)

    filename = f"relatorio_epe_colaborador_{pk}_{dt.date.today().isoformat()}.xlsx"
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


@login_required
def global_report_xlsx(request):
    """
    Exporta atividades do departamento inteiro (com filtros).
    Colunas: Data, Colaborador, Projeto, Geral, Tamanho, Nome do painel, Descrição, Horas(dec), HH:MM:SS
    """
    if not _user_is_manager(request.user):
        return HttpResponseForbidden("Acesso restrito a gestores.")

    # Filtros
    de = (request.GET.get("de") or "").strip()
    ate = (request.GET.get("ate") or "").strip()
    collabs = [v for v in request.GET.getlist("collaborator") if v.isdigit()]
    projects = [v for v in request.GET.getlist("project") if v.isdigit()]
    generals = [v for v in request.GET.getlist("general") if v.isdigit()]
    sizes = [v for v in request.GET.getlist("size") if v.isdigit()]

    qs = (
        Activity.objects.select_related("collaborator", "project", "general", "panel_size", "workday")
        .order_by("collaborator__name", "started_at", "id")
    )
    if de:
        qs = qs.filter(Q(workday__date__gte=de) | Q(workday__isnull=True, started_at__date__gte=de))
    if ate:
        qs = qs.filter(Q(workday__date__lte=ate) | Q(workday__isnull=True, started_at__date__lte=ate))
    if collabs:
        qs = qs.filter(collaborator_id__in=collabs)
    if projects:
        qs = qs.filter(project_id__in=projects)
    if generals:
        qs = qs.filter(general_id__in=generals)
    if sizes:
        qs = qs.filter(panel_size_id__in=sizes)

    # Monta XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Geral"

    headers = [
        "Data",
        "Colaborador",
        "Projeto",
        "Atividade Geral",
        "Tamanho do painel",
        "Nome do painel",
        "Descrição",
        "Horas (dec)",
        "Horas (HH:MM:SS)",
    ]
    ws.append(headers)

    for a in qs:
        secs = a.total_active_seconds()
        hours_dec = round(secs / 3600.0, 2)
        hh = f"{secs//3600:02d}:{(secs%3600)//60:02d}:{secs%60:02d}"
        date_str = a.workday.date.strftime("%Y-%m-%d") if getattr(a, "workday_id", None) else a.started_at.date().isoformat()

        ws.append(
            [
                date_str,
                a.collaborator.name,
                (a.project.name if a.project else ""),
                (a.general.name if a.general_id else ""),
                (a.panel_size.name if a.panel_size_id else ""),
                (a.panel_name or ""),
                (a.description or ""),
                hours_dec,
                hh,
            ]
        )

    # Largura de colunas
    for col in range(1, ws.max_column + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 60)

    filename = f"relatorio_epe_geral_{dt.date.today().isoformat()}.xlsx"
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


# =============================================================================
# Catálogos (índice)
# =============================================================================

@login_required
def catalogs(request):
    generals = GeneralActivity.objects.filter(is_active=True).order_by("order", "name")
    panel_sizes = PanelSize.objects.filter(is_active=True).order_by("order", "name")
    return render(request, "epe/catalogs/index.html", {
        "generals": generals,
        "panel_sizes": panel_sizes,
    })


@login_required
def general_list(request):
    """
    Lista de Atividades Gerais do app EPE.
    Ajuste os order_by/select_related conforme os campos do seu modelo.
    """
    qs = GeneralActivity.objects.filter(is_active=True)

    # Se o seu modelo tiver estes campos, mantenha; senão, simplifique para .order_by("name")
    try:
        qs = qs.select_related("milestone").order_by("milestone__order", "order", "name")
    except Exception:
        qs = qs.order_by("name")

    context = {"objects": qs}
    return render(request, "epe/general_list.html", context)


# epe/views.py
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import PanelSize  # garanta que esse model exista no app epe

@login_required
def size_list(request):
    """
    Lista de Tamanhos de Painel (PanelSize) do app EPE.
    Ajuste order_by conforme os campos do seu modelo.
    """
    qs = PanelSize.objects.all()
    # se seu model tem flags e campos de ordenação, tente ordenar; senão, caia para name
    try:
        qs = qs.filter(is_active=True).order_by("order", "name")
    except Exception:
        qs = qs.order_by("name")

    return render(request, "epe/size_list.html", {"objects": qs})
